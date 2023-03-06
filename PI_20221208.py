# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     PI_20221208
   Description :
   Author :       Jamerri
   date：          2022/12/8
-------------------------------------------------
   Change Activity:
                   2022/12/8:
-------------------------------------------------
"""


import numpy as np
from openpyxl import load_workbook


"""定义参数"""
k_u = 2
k_p = 0.5
PI_1 = []
PI_2 = []
PI_3 = []
PI_4 = []
PI_5 = []
PI_6 = []


def calculate_PI(concentration):
    """PI计算"""
    miu = np.mean(concentration)
    max_value = np.max(concentration)
    pi = k_u * miu + k_p * max_value
    return pi


"""读取数据"""
wb = load_workbook(filename=r"./data/颗粒物-固定传感器采样数据-20221126/无风环境.xlsx")  # 读取路径
sheet = wb["Sheet2"]  # 读取名字为Sheet1的表

Total_time = 300  # 总采样时间
time = 6  # 单次采样时间

for i in range(int(Total_time/time)):
    '''定义空数组'''
    concentration_1 = []
    concentration_2 = []
    concentration_3 = []
    concentration_4 = []
    concentration_5 = []
    concentration_6 = []

    '''分配数据'''
    row_num = 1
    while row_num < 7:
        concentration_1.append(sheet.cell(row=i * time + row_num, column=1).value)
        concentration_2.append(sheet.cell(row=i * time + row_num, column=2).value)
        concentration_3.append(sheet.cell(row=i * time + row_num, column=3).value)
        concentration_4.append(sheet.cell(row=i * time + row_num, column=4).value)
        concentration_5.append(sheet.cell(row=i * time + row_num, column=5).value)
        concentration_6.append(sheet.cell(row=i * time + row_num, column=6).value)
        row_num = row_num + 1

    '''计算PI'''
    PI_1.append(calculate_PI(concentration_1))
    PI_2.append(calculate_PI(concentration_2))
    PI_3.append(calculate_PI(concentration_3))
    PI_4.append(calculate_PI(concentration_4))
    PI_5.append(calculate_PI(concentration_5))
    PI_6.append(calculate_PI(concentration_6))

print(PI_1)
print(PI_2)
print(PI_3)
print(PI_4)
print(PI_5)
print(PI_6)

"""numpy.column_stack() 函数用于将一维数组作为列堆叠到二维数组中"""
raw_data = np.column_stack((PI_1, PI_2, PI_3, PI_4, PI_5, PI_6))

'''写出插值后的数据'''
np.savetxt('./data/颗粒物-固定传感器采样数据-20221126/PI_data.txt', raw_data, fmt="%.2f", delimiter=' ')
