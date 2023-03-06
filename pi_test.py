import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator
from openpyxl import load_workbook


# 读取路径
wb = load_workbook(filename=r"test.xlsx")
# 读取名字为Sheet1的表
sheet = wb["Sheet1"]


concentration = []
concentration_2 = []
pi_counter = []
pi_counter_2 = []
row_num = 1


# PI指标
def PI(index, num, counter):
    if i[0] == i[1] and i[0] == i[2] and i[0] == i[3] and i[0] == i[4] and i[0] == i[5]:
        counter.append(2 * i[0])
    else:
        j = max(i)
        k = i.index(j)
        miu = np.mean(i)
        counter.append(1 * miu + 0.5 * j)


while row_num < 300:
    concentration.append(sheet.cell(row=row_num, column=3).value)
    concentration_2.append(sheet.cell(row=row_num, column=5).value)
    row_num = row_num + 1

split = [concentration[i:i + 6] for i in range(0, len(concentration), 6)]
split2 = [concentration_2[i:i + 6] for i in range(0, len(concentration_2), 6)]
for ind, i in enumerate(split):
    PI(ind, i, pi_counter)

for ind, i in enumerate(split2):
    PI(ind, i, pi_counter_2)
# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=12)

# 设置xtick和ytick的方向：in、out、inout
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

x = np.arange(1, 51)
fig, ax = plt.subplots()

ax.spines['top'].set_visible(False)  # 隐藏上端脊梁
ax.spines['right'].set_visible(False)  # 隐藏上端脊梁

# 设置线条参数
l1, = ax.plot(x, pi_counter, marker='s', markersize=3, color='k', linewidth='1', linestyle='-', clip_on=False)
l2, = ax.plot(x, pi_counter_2, marker='s', markersize=3, color='r', linewidth='1', linestyle='-', clip_on=False)

# 设置图例 bbox_to_anchor图例的位置 ncol设置列数 frameon设置边框
plt.legend(handles=[l1, l2,], labels=['Sensor 0', 'Sensor 1'], bbox_to_anchor=(0.78, 1.1), loc=2, frameon=False)

# 设置坐标轴范围
plt.xlim(0, 50)
ax.set_ylim(0, 100)

ax.set_xlabel('No.', fontsize=12)
ax.set_ylabel('Proximity Index', fontsize=12, color='k')

# 设置X轴主刻度
axxmajorLocator = MultipleLocator(10)
ax.xaxis.set_major_locator(axxmajorLocator)

# 设置Y轴主刻度
axymajorLocator = MultipleLocator(20)
ax.yaxis.set_major_locator(axymajorLocator)

# 保存图片
plt.rcParams['figure.figsize'] = (8.0, 6.0)  # 设置figure_size尺寸
plt.savefig('pi_test.tif', bbox_inches='tight', dpi=500)
plt.show()

