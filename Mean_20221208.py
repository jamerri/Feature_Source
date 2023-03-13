# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     Mean_20221208
   Description :
   Author :       Jamerri
   date：          2022/12/8
-------------------------------------------------
   Change Activity:
                   2022/12/8:
-------------------------------------------------
"""


import numpy as np
from openpyxl import load_workbook


"""定义参数"""
Mean_1 = []
Mean_2 = []
Mean_3 = []
Mean_4 = []
Mean_5 = []
Mean_6 = []


def calculate_mean(concentration):
    """PI计算"""
    mean = np.mean(concentration)
    return mean


"""读取数据"""
wb = load_workbook(filename=r"./data/20230313-1.xlsx")  # 读取路径
sheet = wb["Sheet1"]  # 读取名字为Sheet1的表

Total_time = 360  # 总采样时间
time = 6  # 单次采样时间

for i in range(int(Total_time/time)):
    '''定义空数组'''
    concentration_1 = []
    concentration_2 = []
    concentration_3 = []
    concentration_4 = []
    concentration_5 = []
    concentration_6 = []

    '''分配数据'''
    row_num = 1
    while row_num < 7:
        concentration_1.append(sheet.cell(row=i * time + row_num, column=1).value)
        concentration_2.append(sheet.cell(row=i * time + row_num, column=2).value)
        concentration_3.append(sheet.cell(row=i * time + row_num, column=3).value)
        concentration_4.append(sheet.cell(row=i * time + row_num, column=4).value)
        concentration_5.append(sheet.cell(row=i * time + row_num, column=5).value)
        concentration_6.append(sheet.cell(row=i * time + row_num, column=6).value)
        row_num = row_num + 1

    '''计算Mean'''
    Mean_1.append(calculate_mean(concentration_1))
    Mean_2.append(calculate_mean(concentration_2))
    Mean_3.append(calculate_mean(concentration_3))
    Mean_4.append(calculate_mean(concentration_4))
    Mean_5.append(calculate_mean(concentration_5))
    Mean_6.append(calculate_mean(concentration_6))

print(Mean_1)
print(Mean_2)
print(Mean_3)
print(Mean_4)
print(Mean_5)
print(Mean_6)

"""numpy.column_stack() 函数用于将一维数组作为列堆叠到二维数组中"""
raw_data = np.column_stack((Mean_1, Mean_2, Mean_3, Mean_4, Mean_5, Mean_6))

'''写出插值后的数据'''
np.savetxt('./data/20230313_1_mc.txt', raw_data, fmt="%.2f", delimiter=' ')
