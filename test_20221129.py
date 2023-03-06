# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     test_20221129
   Description :
   Author :       Jamerri
   date：          2022/11/29
-------------------------------------------------
   Change Activity:
                   2022/11/29:
-------------------------------------------------
"""
import numpy as np
import math
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator

'''常数定义'''
"""PI"""
k_u = 2
k_p = 0.5
"""定义Bout_a参数"""
sigma_smooth = 0.3  # sigma值
time_half = 0.4  # 时间半衰期
time_step = 0.1  # 时间步长
"""定义SC参数"""
numda_a = 0.75  # 权重因子
numda_b = 0.25

# 定义数组
Mean_total = []
PI_total = []
SC_total = []
MSC_total = []


# 计算均值的函数
def calculate_mean_value(concentration):
    """均值计算"""
    mean = np.mean(concentration)
    return mean


# 计算最大值的函数
def calculate_max(concentration):
    """P计算"""
    max_value = np.max(concentration)
    # max_no = concentration.index(max_value)
    return max_value


# 计算PI的函数
def calculate_PI(concentration):
    """PI计算"""
    miu = np.mean(concentration)
    max_value = np.max(concentration)
    pi = k_u * miu + k_p * max_value
    return pi


# 计算Bout的函数
def calculate_Bout(concentration):
    """Bout计算"""
    origin_concentration = []
    smooth_concentration = []
    y_concentration = []
    for num in range(len(concentration)):
        origin_concentration.append(concentration[num])
    for num_x in range(len(origin_concentration)):
        s_concentration = origin_concentration[num_x] * math.exp((-1 * 1) / (2 * sigma_smooth * sigma_smooth)) / (sigma_smooth * math.sqrt(2 * math.pi))
        smooth_concentration.append(s_concentration)
    x_concentration = np.diff(smooth_concentration)
    x_concentration = x_concentration.tolist()
    x_concentration.insert(0, 0.0)
    y_concentration.insert(0, 0.0)
    alfa = math.exp(math.log10(1 / (2 * time_half * time_step))) - 1
    for num_y in range(1, 6):
        if num_y == 1:
            y_concentration.append(
                (1 - alfa) * y_concentration[0] + alfa * (x_concentration[num_y] - x_concentration[0]))
        else:
            y_concentration.append(
                (1 - alfa) * y_concentration[num_y - 1] + alfa * (x_concentration[num_y] - x_concentration[num_y - 1]))
    y_derivative = np.diff(y_concentration)
    y_derivative = y_derivative.tolist()
    bout = 0
    aa = []
    for num_z in range(0, 5):
        if y_derivative[num_z] >= 0:
            aa.append(1)
        else:
            aa.append(-1)
    for num_h in range(1, 4):
        if aa[num_h] - aa[num_h - 1] == 0 and aa[num_h + 1] - aa[num_h] == -2:
            bout += 1
        else:
            bout = bout
    return bout


def calculate_Bout_amp(concentration):
    """Bout计算"""
    origin_concentration = []
    smooth_concentration = []
    y_concentration = []
    Bout_amp = 0
    for num in range(len(concentration)):
        origin_concentration.append(concentration[num])
    for num_x in range(len(origin_concentration)):
        s_concentration = origin_concentration[num_x] * math.exp((-1 * 1) / (2 * sigma_smooth * sigma_smooth)) / (sigma_smooth * math.sqrt(2 * math.pi))
        smooth_concentration.append(s_concentration)
    x_concentration = np.diff(smooth_concentration)
    x_concentration = x_concentration.tolist()
    x_concentration.insert(0, 0.0)
    y_concentration.insert(0, 0.01)
    alfa = math.exp(math.log10(1 / (2 * time_half * time_step))) - 1
    for num_y in range(1, 6):
        if num_y == 1:
            y_concentration.append(
                (1 - alfa) * y_concentration[0] + alfa * (x_concentration[num_y] - x_concentration[0]))
        else:
            y_concentration.append(
                (1 - alfa) * y_concentration[num_y - 1] + alfa * (x_concentration[num_y] - x_concentration[num_y - 1]))
    y_derivative = np.diff(y_concentration)
    pos_change = []
    neg_change = []

    for i in range(len(y_derivative)):
        if y_derivative[i] > 0:
            pos_change.append(i)
        if y_derivative[i] < 0:
            neg_change.append(i)

    if len(pos_change) == 0 or len(neg_change) == 0 or (pos_change[0] > neg_change[0] and len(neg_change) == 1):
        Bout_amp = 0.0
    else:
        if pos_change[0] > neg_change[0]:
            del (neg_change[0])

        if len(pos_change) > len(neg_change):
            difference = len(pos_change) - len(neg_change)
            pos_change = pos_change[:-difference]

        posneg = [[0] * (len(pos_change) + 1) for i in range(3)]
        for i in range(len(pos_change)):
            posneg[0][i] = pos_change[i]
            posneg[1][i] = neg_change[i]

        amps = [0 for i in range(len(pos_change))]

        for i in range(len(pos_change)):
            amps[i] = y_concentration[posneg[1][i]] - y_concentration[posneg[0][i]]

        sum = 0
        for i in range(len(amps)):
            sum += amps[i]

        amps_averge = sum / len(amps)
        Bout_amp = amps_averge
    return Bout_amp


# 读取数据
wb = load_workbook(filename=r"./data/颗粒物-固定传感器-20221126.xlsx")  # 读取路径
sheet = wb["Sheet1"]  # 读取名字为Sheet1的表

Total_time = 300  # 总采样时间
time = 6 # 窗口时间

for i in range(int(Total_time/time)):
    """定义空数组"""
    concentration_1 = []
    concentration_2 = []
    concentration_3 = []
    concentration_4 = []
    concentration_5 = []
    concentration_6 = []

    Mean = []
    PI = []
    SC = []
    MSC = []

    """分段数据"""
    row_num = 1
    while row_num < (time + 1):
        concentration_1.append(sheet.cell(row=row_num + i * time, column=1).value)
        concentration_2.append(sheet.cell(row=row_num + i * time, column=2).value)
        concentration_3.append(sheet.cell(row=row_num + i * time, column=3).value)
        concentration_4.append(sheet.cell(row=row_num + i * time, column=4).value)
        concentration_5.append(sheet.cell(row=row_num + i * time, column=5).value)
        concentration_6.append(sheet.cell(row=row_num + i * time, column=6).value)
        row_num = row_num + 1

    Mean_1 = calculate_mean_value(concentration_1)
    Mean_2 = calculate_mean_value(concentration_2)
    Mean_3 = calculate_mean_value(concentration_3)
    Mean_4 = calculate_mean_value(concentration_4)
    Mean_5 = calculate_mean_value(concentration_5)
    Mean_6 = calculate_mean_value(concentration_6)

    Mean.append(Mean_1)
    Mean.append(Mean_2)
    Mean.append(Mean_3)
    Mean.append(Mean_4)
    Mean.append(Mean_5)
    Mean.append(Mean_6)

    Mean_total.append(Mean)

    PI_1 = calculate_PI(concentration_1)
    PI_2 = calculate_PI(concentration_2)
    PI_3 = calculate_PI(concentration_3)
    PI_4 = calculate_PI(concentration_4)
    PI_5 = calculate_PI(concentration_5)
    PI_6 = calculate_PI(concentration_6)

    PI.append(PI_1)
    PI.append(PI_2)
    PI.append(PI_3)
    PI.append(PI_4)
    PI.append(PI_5)
    PI.append(PI_6)

    PI_total.append(PI)

    SC_1 = numda_a * calculate_mean_value(concentration_1) + numda_b * calculate_Bout(concentration_1) * calculate_max(concentration_1)
    SC_2 = numda_a * calculate_mean_value(concentration_2) + numda_b * calculate_Bout(concentration_2) * calculate_max(concentration_2)
    SC_3 = numda_a * calculate_mean_value(concentration_3) + numda_b * calculate_Bout(concentration_3) * calculate_max(concentration_3)
    SC_4 = numda_a * calculate_mean_value(concentration_4) + numda_b * calculate_Bout(concentration_4) * calculate_max(concentration_4)
    SC_5 = numda_a * calculate_mean_value(concentration_5) + numda_b * calculate_Bout(concentration_5) * calculate_max(concentration_5)
    SC_6 = numda_a * calculate_mean_value(concentration_6) + numda_b *  calculate_Bout(concentration_6) * calculate_max(concentration_6)

    SC.append(SC_1)
    SC.append(SC_2)
    SC.append(SC_3)
    SC.append(SC_4)
    SC.append(SC_5)
    SC.append(SC_6)

    SC_total.append(SC)

    MSC_1 = numda_a * calculate_mean_value(concentration_1) + numda_b * calculate_Bout_amp(concentration_1) * time
    MSC_2 = numda_a * calculate_mean_value(concentration_2) + numda_b * calculate_Bout_amp(concentration_2) * time
    MSC_3 = numda_a * calculate_mean_value(concentration_3) + numda_b * calculate_Bout_amp(concentration_3) * time
    MSC_4 = numda_a * calculate_mean_value(concentration_4) + numda_b * calculate_Bout_amp(concentration_4) * time
    MSC_5 = numda_a * calculate_mean_value(concentration_5) + numda_b * calculate_Bout_amp(concentration_5) * time
    MSC_6 = numda_a * calculate_mean_value(concentration_6) + numda_b * calculate_Bout_amp(concentration_6) * time

    MSC.append(MSC_1)
    MSC.append(MSC_2)
    MSC.append(MSC_3)
    MSC.append(MSC_4)
    MSC.append(MSC_5)
    MSC.append(MSC_6)

    MSC_total.append(MSC)

print(Mean_total)
print(PI_total)
print(SC_total)
print(MSC_total)


def drawing_function_1(mean, pi, sc, msc, num):
    # 导入Times New Roman字体
    plt.rc('font', family='Times New Roman', size=12)

    # 设置xtick和ytick的方向：in、out、inout
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'

    x = np.arange(0, 50)
    # x = [1, 2, 3, 4, 5, 6]
    fig, ax = plt.subplots()

    ax.spines['top'].set_visible(False)  # 隐藏上端脊梁
    ax.spines['right'].set_visible(False)  # 隐藏上端脊梁

    # 设置线条参数
    l1, = ax.plot(x, mean, marker='s', markersize=3, color='k', linewidth='1', linestyle='-', clip_on=False)
    l2, = ax.plot(x, pi, marker='s', markersize=3, color='r', linewidth='1', linestyle='-', clip_on=False)
    l3, = ax.plot(x, sc, marker='s', markersize=3, color='g', linewidth='1', linestyle='-', clip_on=False)
    l4, = ax.plot(x, msc, marker='s', markersize=3, color='k', linewidth='1', linestyle='--', clip_on=False)

    ax.set_xlabel('No.', fontsize=12)
    ax.set_ylabel('MSC', fontsize=12, color='k')


    # 保存图片
    plt.rcParams['figure.figsize'] = (8.0, 6.0)  # 设置figure_size尺寸
    # plt.savefig('./data/test-' + str(num) + '-.tif', bbox_inches='tight', dpi=500)
    plt.show()


# for i in range(len(Mean_total)):
#     Mean_value = Mean_total[i]
#     PI_value = PI_total[i]
#     SC_value = SC_total[i]
#     MSC_value = MSC_total[i]
#     drawing_function_1(Mean_value, PI_value, SC_value, MSC_value, i)
#     print("ok !!!")


def drawing_function_2(value):
    # 导入Times New Roman字体
    plt.rc('font', family='Times New Roman', size=12)

    # 设置xtick和ytick的方向：in、out、inout
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'

    # x = np.arange(0, 50)
    x = [1, 2, 3, 4, 5, 6]
    fig, ax = plt.subplots()

    ax.spines['top'].set_visible(False)  # 隐藏上端脊梁
    ax.spines['right'].set_visible(False)  # 隐藏上端脊梁

    # 设置线条参数
    l1, = ax.plot(x, value[0], marker='s', markersize=3, color='k', linewidth='1', linestyle='-', clip_on=False)
    l2, = ax.plot(x, value[1], marker='s', markersize=3, color='r', linewidth='1', linestyle='-', clip_on=False)
    l3, = ax.plot(x, value[2], marker='s', markersize=3, color='g', linewidth='1', linestyle='-', clip_on=False)
    l4, = ax.plot(x, value[3], marker='s', markersize=3, color='k', linewidth='1', linestyle='--', clip_on=False)
    l5, = ax.plot(x, value[4], marker='s', markersize=3, color='r', linewidth='1', linestyle='--', clip_on=False)
    l6, = ax.plot(x, value[5], marker='s', markersize=3, color='g', linewidth='1', linestyle='--', clip_on=False)

    ax.set_xlabel('No.', fontsize=12)
    ax.set_ylabel('MSC', fontsize=12, color='k')


    # 保存图片
    plt.rcParams['figure.figsize'] = (8.0, 6.0)  # 设置figure_size尺寸
    # plt.savefig('./data/Mean.tif', bbox_inches='tight', dpi=500)
    plt.show()


Value = MSC_total
drawing_function_2(Value)

Value = SC_total
drawing_function_2(Value)

Value = PI_total
drawing_function_2(Value)

Value = Mean_total
drawing_function_2(Value)