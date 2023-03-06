# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     SC_20221031
   Description :
   Author :       Jamerri
   date：          2022/10/31
-------------------------------------------------
   Change Activity:
                   2022/10/31:
-------------------------------------------------
"""


import numpy as np
import math
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator


"""定义PI参数"""
k_u = 2
k_p = 0.5
PI_1 = []
PI_2 = []
PI_3 = []

"""定义Bout_a参数"""
sigma_smooth: float = 0.3  # sigma值
time_half = 0.4  # 时间半衰期
time_step = 0.1  # 时间步长

Bout_1 = []
Bout_2 = []
Bout_3 = []

"""定义SC_New参数"""
numda = 0.75  # 权重因子

SC_1 = []
SC_2 = []
SC_3 = []


def calculate_mean_value(concentration):
    mean = np.mean(concentration)
    return mean


def calculate_P(concentration):
    """P计算"""
    max_value = np.max(concentration)
    # max_no = concentration.index(max_value)
    return max_value


def calculate_Bout(concentration):
    """Bout计算"""
    origin_concentration = []
    smooth_concentration = []
    y_concentration = []
    for num in range(len(concentration)):
        origin_concentration.append(concentration[num])
    for num_x in range(len(origin_concentration)):
        s_concentration = origin_concentration[num_x] * math.exp((-1 * 1) / (2 * sigma_smooth * sigma_smooth)) / (sigma_smooth * math.sqrt(2 * math.pi))
        smooth_concentration.append(s_concentration)
    x_concentration = np.diff(smooth_concentration)
    x_concentration = x_concentration.tolist()
    x_concentration.insert(0, 0.0)
    y_concentration.insert(0, 0.0)
    alfa = math.exp(math.log10(1 / (2 * time_half * time_step))) - 1
    for num_y in range(1, 6):
        if num_y == 1:
            y_concentration.append(
                (1 - alfa) * y_concentration[0] + alfa * (x_concentration[num_y] - x_concentration[0]))
        else:
            y_concentration.append(
                (1 - alfa) * y_concentration[num_y - 1] + alfa * (x_concentration[num_y] - x_concentration[num_y - 1]))
    y_derivative = np.diff(y_concentration)
    y_derivative = y_derivative.tolist()
    bout = 0
    aa = []
    for num_z in range(0, 5):
        if y_derivative[num_z] >= 0:
            aa.append(1)
        else:
            aa.append(-1)
    for num_h in range(1, 4):
        if aa[num_h] - aa[num_h - 1] == 0 and aa[num_h + 1] - aa[num_h] == -2:
            bout += 1
        else:
            bout = bout
    return bout


"""读取数据"""
wb = load_workbook(filename=r"test.xlsx")  # 读取路径
sheet = wb["Sheet1"]  # 读取名字为Sheet1的表

Total_time = 300  # 总采样时间
time = 6  # 单次采样时间

for i in range(int(Total_time/time)):
    '''定义空数组'''
    concentration_1 = []
    concentration_2 = []
    concentration_3 = []

    '''分配数据'''
    row_num = 1
    while row_num < 7:
        concentration_1.append(sheet.cell(row=row_num * i + 1, column=1).value / 4)
        concentration_2.append(sheet.cell(row=row_num * i + 1, column=2).value / 4)
        concentration_3.append(sheet.cell(row=row_num * i + 1, column=3).value / 4)
        row_num = row_num + 1

    sensor_1 = numda * calculate_mean_value(concentration_1) + (1 - numda) * calculate_P(
        concentration_1) * calculate_Bout(
        concentration_1)

    sensor_2 = numda * calculate_mean_value(concentration_2) + (1 - numda) * calculate_P(
        concentration_2) * calculate_Bout(
        concentration_2)

    sensor_3 = numda * calculate_mean_value(concentration_3) + (1 - numda) * calculate_P(
        concentration_3) * calculate_Bout(
        concentration_3)

    SC_1.append(sensor_1)
    SC_2.append(sensor_2)
    SC_3.append(sensor_3)


print(SC_1)
print(SC_2)
print(SC_3)

# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=12)

# 设置xtick和ytick的方向：in、out、inout
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

x = np.arange(0, 50)
fig, ax = plt.subplots()

ax.spines['top'].set_visible(False)  # 隐藏上端脊梁
ax.spines['right'].set_visible(False)  # 隐藏上端脊梁

# 设置线条参数
l1, = ax.plot(x, SC_1, marker='s', markersize=3, color='k', linewidth='1', linestyle='-', clip_on=False)
l2, = ax.plot(x, SC_2, marker='s', markersize=3, color='r', linewidth='1', linestyle='-', clip_on=False)
l3, = ax.plot(x, SC_3, marker='s', markersize=3, color='g', linewidth='1', linestyle='-', clip_on=False)

# 设置图例 bbox_to_anchor图例的位置 ncol设置列数 frameon设置边框
plt.legend(handles=[l1, l2, l3], labels=['Sensor 0', 'Sensor 1', 'Sensor 2'], bbox_to_anchor=(0.78, 1.1), loc=2, frameon=False)

# 设置坐标轴范围
plt.xlim(0, 50)
ax.set_ylim(0, 50)

ax.set_xlabel('No.', fontsize=12)
ax.set_ylabel('Sc', fontsize=12, color='k')

# 设置X轴主刻度
axxmajorLocator = MultipleLocator(10)
ax.xaxis.set_major_locator(axxmajorLocator)

# 设置Y轴主刻度
axymajorLocator = MultipleLocator(10)
ax.yaxis.set_major_locator(axymajorLocator)

# 保存图片
plt.rcParams['figure.figsize'] = (8.0, 6.0)  # 设置figure_size尺寸
# plt.savefig('pi_test.tif', bbox_inches='tight', dpi=500)
plt.show()