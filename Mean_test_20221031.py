# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     Mean_test_20221031
   Description :
   Author :       Jamerri
   date：          2022/10/31
-------------------------------------------------
   Change Activity:
                   2022/10/31:
-------------------------------------------------
"""


import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator


"""定义参数"""
k_u = 2
k_p = 0.5
P_1 = []
P_2 = []
P_3 = []


def calculate_P(concentration):
    """PI计算"""
    mean = np.mean(concentration)
    return mean


"""读取数据"""
wb = load_workbook(filename=r"test.xlsx")  # 读取路径
sheet = wb["Sheet1"]  # 读取名字为Sheet1的表

Total_time = 300  # 总采样时间
time = 6  # 单次采样时间

for i in range(int(Total_time/time)):
    '''定义空数组'''
    concentration_1 = []
    concentration_2 = []
    concentration_3 = []

    '''分配数据'''
    row_num = 1
    while row_num < 7:
        concentration_1.append(sheet.cell(row=row_num * i + 1, column=1).value / 4)
        concentration_2.append(sheet.cell(row=row_num * i + 1, column=2).value / 4)
        concentration_3.append(sheet.cell(row=row_num * i + 1, column=3).value / 4)
        row_num = row_num + 1

    '''计算PI'''
    P_1.append(calculate_P(concentration_1))
    P_2.append(calculate_P(concentration_2))
    P_3.append(calculate_P(concentration_3))


print(P_1)
print(P_2)
print(P_3)

# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=12)

# 设置xtick和ytick的方向：in、out、inout
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'

x = np.arange(0, 50)
fig, ax = plt.subplots()

ax.spines['top'].set_visible(False)  # 隐藏上端脊梁
ax.spines['right'].set_visible(False)  # 隐藏上端脊梁

# 设置线条参数
l1, = ax.plot(x, P_1, marker='s', markersize=3, color='k', linewidth='1', linestyle='-', clip_on=False)
l2, = ax.plot(x, P_2, marker='s', markersize=3, color='r', linewidth='1', linestyle='-', clip_on=False)
l3, = ax.plot(x, P_3, marker='s', markersize=3, color='g', linewidth='1', linestyle='-', clip_on=False)

# 设置图例 bbox_to_anchor图例的位置 ncol设置列数 frameon设置边框
plt.legend(handles=[l1, l2, l3], labels=['Sensor 0', 'Sensor 1', 'Sensor 2'], bbox_to_anchor=(0.78, 1.1), loc=2, frameon=False)

# 设置坐标轴范围
plt.xlim(0, 50)
ax.set_ylim(0, 50)

ax.set_xlabel('No.', fontsize=12)
ax.set_ylabel('Mean value', fontsize=12, color='k')

# 设置X轴主刻度
axxmajorLocator = MultipleLocator(10)
ax.xaxis.set_major_locator(axxmajorLocator)

# 设置Y轴主刻度
axymajorLocator = MultipleLocator(10)
ax.yaxis.set_major_locator(axymajorLocator)

# 保存图片
plt.rcParams['figure.figsize'] = (8.0, 6.0)  # 设置figure_size尺寸
# plt.savefig('pi_test.tif', bbox_inches='tight', dpi=500)
plt.show()