# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     test_20221126
   Description :
   Author :       Jamerri
   date：          2022/11/26
-------------------------------------------------
   Change Activity:
                   2022/11/26:
-------------------------------------------------
"""
import random
from scipy.interpolate import griddata
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import matplotlib.colors as col
import matplotlib.cm as cm


# 自定义参数
x_origin = []
y_origin = []
c = []

# 源位置参数
X_position = 6.85
Y_position = 2.25

# 读取数据
book = load_workbook(filename=r"./data/20221216-6.85-2.25/pm_robot.xlsx")  # 文件路径
sheet = book["Sheet1"]  # 读取名字为Sheet1的表

row_num = 2
while row_num <= 33:
    '''将表中第 i 列的2-33行数据写入数组中'''
    x_origin.append('%.2f' % sheet.cell(row=row_num, column=1).value)
    row_num = row_num + 1

row_num = 2
while row_num <= 33:
    '''将表中第 i 列的2-33行数据写入数组中'''
    y_origin.append('%.2f' % sheet.cell(row=row_num, column=2).value)
    row_num = row_num + 1

row_num = 2
while row_num <= 33:
    '''将表中第 i 列的2-33行数据写入数组中'''
    c.append('%.2f' % sheet.cell(row=row_num, column=4).value)
    row_num = row_num + 1

# print(x_origin)
# print(y_origin)

# 坐标轴转换为浮点数
x_position = []
y_position = []
c_value = []
for num_x in range(len(x_origin)):
    x_position.append(float(x_origin[num_x]))
for num_y in range(len(y_origin)):
    y_position.append(float(y_origin[num_y]))
for num_c in range(len(c)):
    c_value.append(float(c[num_c]))

print(x_position)
print(y_position)
print(c_value)

raw_data = np.column_stack((x_position, y_position, c_value))  # numpy.column_stack() 函数用于将一维数组作为列堆叠到二维数组中
# print(raw_data)

# 生成空网格
'''定义参数'''
x_min = 0.5
x_max = 7.5
y_min = 0.5
y_max = 3.5
x_cell_size = 0.25
y_cell_size = 0.25
x = np.arange(x_min, x_max + x_cell_size, x_cell_size)
print("x方向网格数量为:", len(x))
y = np.arange(y_min, y_max + y_cell_size, y_cell_size)
print("y方向网格数量为:", len(y))
print("网格总数为:", len(x)*len(y))
empty_grid = []
for i in y:
    for j in x:
        empty_grid.append([round(j, 2), round(i, 2)])
empty_grid = np.array(empty_grid)
print("-------------------empty_grid----------------------------")
# print(empty_grid)

'''进行插值，把原始数据集插值到空网格坐标节点上'''
grid_c = griddata(raw_data[:, :2], raw_data[:, 2], empty_grid, method='linear')  # griddata()函数实现二维插值。
print(raw_data[:, :2])
print(raw_data[:, 2])
interpolation_data = np.column_stack((empty_grid, grid_c))  # 合并
print(interpolation_data)

interpolation_x = interpolation_data[:, 0]  # 取出x值
interpolation_y = interpolation_data[:, 1]  # 取出y值
interpolation_c = interpolation_data[:, 2]  # 取出浓度插值后的值
'''归一化'''
c_min = min(interpolation_c)
c_max = max(interpolation_c)
interpolation_c = (np.array(interpolation_c)-c_min)/(c_max-c_min)
for i in range(len(interpolation_c)):
    if interpolation_c[i] < 0:
        interpolation_c[i] = 0
interpolation_x_new = np.array(interpolation_x).reshape(len(y), len(x))
interpolation_y_new = np.array(interpolation_y).reshape(len(y), len(x))
interpolation_c_new = np.array(interpolation_c).reshape(len(y), len(x))
print(interpolation_c_new)

'''写出插值后的数据'''
np.savetxt('./data/20221216-6.85-2.25/pi_interpolation_data.txt', interpolation_data, fmt="%.2f", delimiter=' ')


'''颜色bar调整'''
color_1 = '#ff0000'  # 红色
color_2 = '#00ff00'  # 绿色
color_3 = '#0000CD'  # 蓝色
color_4 = '#ffff00'  # 黄色
color_5 = '#00bfff'  # 深天蓝
color_6 = '#4169E1'  # 宝蓝色
color_7 = '#1E90FF'  # 闪蓝色
color_8 = '#6495ED'  # 矢车菊色
color_9 = '#00FFFF'  # 青色
color_10 = '#FF9900'  # 橙色
color_11 = '#ff8C00'  # 橙红色
color_12 = '#87CEFA'  # 浅天蓝

cmap = col.LinearSegmentedColormap.from_list('own', [color_3,  color_9, color_2, color_4,  color_1])
cm.register_cmap(cmap=cmap)
cm.get_cmap('own')

# 导入Times New Roman字体
plt.rc('font', family='Times New Roman', size=12)

# 设置坐标轴标题
plt.xlabel('X (m)', fontsize=12)  # x轴标题
plt.ylabel('Y (m)', fontsize=12)  # y轴标题

# 设置坐标轴范围
plt.xlim(0, 8)
plt.ylim(0, 4.1)

# 设置坐标轴单位长度一致
ax = plt.gca()
ax.set_aspect(1)

# 源位置图案
plt.plot(X_position, Y_position, 'k-', marker='o', markersize=5)

# 画热力图
plt.rcParams['figure.figsize'] = (12.0, 6.0)  # 设置figure_size尺寸
plt.pcolormesh(interpolation_x_new, interpolation_y_new, interpolation_c_new, cmap='own', vmin=0, vmax=1)
# plt.pcolormesh(interpolation_x_new, interpolation_y_new, interpolation_c_new, edgecolors='black', linewidths=0.05, cmap='own', vmin=0, vmax=1)
# plt.imshow(interpolation_c_new, cmap='own', vmin=0, vmax=1, origin='lower')
plt.axis('off')
plt.colorbar()

plt.title("MPI", fontdict=None, loc="center")
# 保存图片
plt.savefig('./data/20221216-6.85-2.25/pi.tif', bbox_inches='tight', dpi=600)
plt.show()
