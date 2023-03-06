# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     Mean_20221126
   Description :
   Author :       Jamerri
   date：          2022/11/26
-------------------------------------------------
   Change Activity:
                   2022/11/26:
-------------------------------------------------
"""


import numpy as np
from openpyxl import load_workbook


"""定义参数"""
P_1 = []


def calculate_P(concentration):
    """PI计算"""
    mean = np.mean(concentration)
    return mean


"""读取数据"""
wb = load_workbook(filename=r"./data/移动机器人采样-第四组-10hz-7.15-2.05/pm_cocnetration.xlsx")  # 读取路径
sheet = wb["Sheet1"]  # 读取名字为Sheet1的表

Total_time = 1920  # 总采样时间
time = 60  # 单次采样时间

for i in range(int(Total_time/time)):
    '''定义空数组'''
    concentration_1 = []

    '''分配数据'''
    row_num = 1
    while row_num < 61:
        concentration_1.append(sheet.cell(row=i * time + row_num, column=1).value)
        row_num = row_num + 1

    '''计算PI'''
    P_1.append(calculate_P(concentration_1))

print(P_1)

'''写出插值后的数据'''
np.savetxt('./data/移动机器人采样-第四组-10hz-7.15-2.05/Mean_data.txt', P_1, fmt="%.2f", delimiter=' ')
