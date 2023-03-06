# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     PI_20221126
   Description :
   Author :       Jamerri
   date：          2022/11/26
-------------------------------------------------
   Change Activity:
                   2022/11/26:
-------------------------------------------------
"""


import numpy as np
from openpyxl import load_workbook


"""定义参数"""
k_u = 2
k_p = 0.5
PI_1 = []


def calculate_PI(concentration):
    """PI计算"""
    miu = np.mean(concentration)
    max_value = np.max(concentration)
    pi = k_u * miu + k_p * max_value
    return pi


"""读取数据"""
wb = load_workbook(filename=r"./data/移动机器人采样-第四组-10hz-7.15-2.05/pm_cocnetration.xlsx")  # 读取路径
sheet = wb["Sheet1"]  # 读取名字为Sheet1的表

Total_time = 1920  # 总采样时间
time = 60  # 单次采样时间

for i in range(int(Total_time/time)):
    '''定义空数组'''
    concentration_1 = []

    '''分配数据'''
    row_num = 1
    while row_num < 61:
        concentration_1.append(sheet.cell(row=i * time + row_num, column=1).value)
        row_num = row_num + 1

    '''计算PI'''
    PI_1.append(calculate_PI(concentration_1))

print(PI_1)

'''写出插值后的数据'''
np.savetxt('./data/移动机器人采样-第四组-10hz-7.15-2.05/PI_data.txt', PI_1, fmt="%.2f", delimiter=' ')