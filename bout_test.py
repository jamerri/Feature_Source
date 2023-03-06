#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time : 2022/4/22 10:34 上午
# @Author : Jamerri
# @FileName: bout_test.py
# @Email : jamerri@163.com
# @Software: PyCharm

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator
from openpyxl import load_workbook
import math


# 读取路径
wb = load_workbook(filename=r"test.xlsx")
# 读取名字为Sheet1的表
sheet = wb["Sheet1"]


concentration = []
total_s_concentration = []
bout_counter = []
interval = 60  # 采样间隔
row_num = 1
s_sigma = 0.3  # sigma值
time_half = 0.4  # 时间半衰期
time_step = 0.1  # 时间步长
count = 1  # 计数器

# 读取浓度数据
while row_num < 301:
    concentration.append(sheet.cell(row=row_num, column=3).value)
    row_num = row_num + 1

# 将浓度数据转换为列表
concentration = list(map(float, concentration))

# 按照间隔分割数据
split = [concentration[i:i + interval] for i in range(0, len(concentration), interval)]


# 每个间隔原始浓度和y浓度图像制作函数
def make_figure(y_c, origin_c, count):
    # 导入Times New Roman字体
    plt.rc('font', family='Times New Roman', size=12)

    # 设置xtick和ytick的方向：in、out、inout
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'

    # X数据（时间Time）
    x = np.arange(0, interval, 1)

    fig, ax1 = plt.subplots()
    ax1.spines['top'].set_visible(False)  # 隐藏上端脊梁
    ax2 = ax1.twinx()  # 共享X轴
    ax2.spines['top'].set_visible(False)  # 隐藏上端脊梁

    # 设置线条参数
    l1, = ax1.plot(x, y_c, 'g-')
    l2, = ax2.plot(x, origin_c, 'b-')

    # 设置坐标轴范围
    plt.xlim(0, interval)

    # 设置坐标轴标签
    ax1.set_xlabel('Time (s)', fontsize=12)
    ax1.set_ylabel('Smooth Concentration', fontsize=12, color='k')
    ax2.set_ylabel('Inital Concentration', fontsize=12, color='k')

    # 设置图例 bbox_to_anchor图例的位置 ncol设置列数 frameon设置边框
    # plt.legend(handles=[l1, l2, ], labels=['Smooth Concentration', 'Inital Concentration'], bbox_to_anchor=(0.18, 1.1),
    #            loc=2, ncol=2, frameon=False)
    plt.rcParams['figure.figsize'] = (4, 3)  # 设置figure_size尺寸
    save_path = 'No.' + str(count) + '-interval.tiff'
    plt.savefig(save_path, bbox_inches='tight', dpi=300)  # 保存图片
    plt.show()


# bout数曲线图像制作函数
def boot_make_figure(b_counter, count_number):
    # 导入Times New Roman字体
    plt.rc('font', family='Times New Roman', size=12)

    # 设置xtick和ytick的方向：in、out、inout
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'

    x = np.arange(1, count_number + 1)
    fig, ax = plt.subplots()

    ax.spines['top'].set_visible(False)  # 隐藏上端脊梁
    ax.spines['right'].set_visible(False)  # 隐藏上端脊梁

    # 设置线条参数
    l1, = ax.plot(x, b_counter, marker='s', markersize=3, color='k', linewidth='1', linestyle='-', clip_on=False)

    # 设置图例 bbox_to_anchor图例的位置 ncol设置列数 frameon设置边框
    # plt.legend(handles=[l1, ], labels=['Sensor 0', 'Sensor 1'], bbox_to_anchor=(0.78, 1.1), loc=2, frameon=False)

    # 设置X轴主刻度
    axxmajorLocator = MultipleLocator(1)
    ax.xaxis.set_major_locator(axxmajorLocator)

    ax.set_xlabel('No.', fontsize=12)
    ax.set_ylabel('Bout numbers', fontsize=12, color='k')

    # 保存图片
    plt.rcParams['figure.figsize'] = (4, 3)  # 设置figure_size尺寸
    # plt.savefig('bout.tiff', bbox_inches='tight', dpi=300)
    plt.show()


# 计算Bout指标
for ind, i in enumerate(split):
    origin_concentration = []
    smooth_concentration = []
    x_concentration = []
    y_concentration = []
    for num_x in range(0, interval):
        origin_concentration.append(float(i[num_x]))
        s_concentration = origin_concentration[num_x] * math.exp((-1 * 1) / (2 * s_sigma * s_sigma)) / (s_sigma * math.sqrt(2 * math.pi))
        smooth_concentration.append(s_concentration)
    x_concentration = np.diff(smooth_concentration)
    print(len(x_concentration))
    x_concentration = x_concentration.tolist()
    x_concentration.insert(0, 0.0)
    y_concentration.insert(0, 0.0)
    alfa = math.exp(math.log10(1 / (2 * time_half * time_step))) - 1
    # alfa = 1 - math.exp(math.log10(0.5)/ time_half)
    # alfa = 1 - 0.5 ** (1 / time_half * time_step)
    for num_y in range(1, interval):
        print(type(y_concentration[0]))
        print(type(x_concentration[num_y]))
        if num_y == 1:
            y_concentration.append((1 - alfa) * y_concentration[0] + alfa * (x_concentration[num_y] - x_concentration[0]))
        else:
            y_concentration.append((1 - alfa) * y_concentration[num_y - 1] + alfa * (x_concentration[num_y] - x_concentration[num_y - 1]))
    y_derivative = np.diff(y_concentration)
    y_derivative = y_derivative.tolist()
    make_figure(y_concentration, origin_concentration, count)
    bout = 0
    for num_z in range(0, interval-1):
        if y_derivative[num_z] > 0:
            bout += 1
        else:
            bout = bout
    bout_counter.append(bout)
    count += 1
boot_make_figure(bout_counter, len(bout_counter))
