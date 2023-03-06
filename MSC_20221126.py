# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     MSC_20221126
   Description :
   Author :       Jamerri
   date：          2022/11/26
-------------------------------------------------
   Change Activity:
                   2022/11/26:
-------------------------------------------------
"""


import numpy as np
import math
from openpyxl import load_workbook

"""定义PI参数"""
k_u = 2
k_p = 0.5
PI_1 = []

"""定义Bout_a参数"""
sigma_smooth: float = 0.3  # sigma值
time_half = 0.4  # 时间半衰期
time_step = 0.1  # 时间步长

Bout_1 = []

"""定义SC_New参数"""
numda = 0.5  # 权重因子

SC_1 = []


def calculate_mean_value(concentration):
    mean = np.mean(concentration)
    return mean


def calculate_P(concentration):
    """P计算"""
    max_value = np.max(concentration)
    # max_no = concentration.index(max_value)
    return max_value


def calculate_Bout(concentration):
    """Bout计算"""
    origin_concentration = []
    smooth_concentration = []
    y_concentration = []
    Bout_amp = 0
    for num in range(len(concentration)):
        origin_concentration.append(concentration[num])
    for num_x in range(len(origin_concentration)):
        s_concentration = origin_concentration[num_x] * math.exp((-1 * 1) / (2 * sigma_smooth * sigma_smooth)) / (sigma_smooth * math.sqrt(2 * math.pi))
        smooth_concentration.append(s_concentration)
    x_concentration = np.diff(smooth_concentration)
    x_concentration = x_concentration.tolist()
    x_concentration.insert(0, 0.0)
    y_concentration.insert(0, 0.01)
    alfa = math.exp(math.log10(1 / (2 * time_half * time_step))) - 1
    for num_y in range(1, 60):
        if num_y == 1:
            y_concentration.append(
                (1 - alfa) * y_concentration[0] + alfa * (x_concentration[num_y] - x_concentration[0]))
        else:
            y_concentration.append(
                (1 - alfa) * y_concentration[num_y - 1] + alfa * (x_concentration[num_y] - x_concentration[num_y - 1]))
    y_derivative = np.diff(y_concentration)
    pos_change = []
    neg_change = []

    for i in range(len(y_derivative)):
        if y_derivative[i] > 0:
            pos_change.append(i)
        if y_derivative[i] < 0:
            neg_change.append(i)

    if len(pos_change) == 0 or len(neg_change) == 0 or (pos_change[0] > neg_change[0] and len(neg_change) == 1):
        Bout_amp = 0.0
    else:
        if pos_change[0] > neg_change[0]:
            del (neg_change[0])

        if len(pos_change) > len(neg_change):
            difference = len(pos_change) - len(neg_change)
            pos_change = pos_change[:-difference]

        posneg = [[0] * (len(pos_change) + 1) for i in range(3)]
        for i in range(len(pos_change)):
            posneg[0][i] = pos_change[i]
            posneg[1][i] = neg_change[i]

        amps = [0 for i in range(len(pos_change))]

        for i in range(len(pos_change)):
            amps[i] = y_concentration[posneg[1][i]] - y_concentration[posneg[0][i]]

        sum = 0
        for i in range(len(amps)):
            sum += amps[i]

        amps_averge = sum / len(amps)
        Bout_amp = amps_averge
    return Bout_amp


"""读取数据"""
wb = load_workbook(filename=r"./data/移动机器人采样-第四组-10hz-7.15-2.05/pm_cocnetration.xlsx")  # 读取路径
sheet = wb["Sheet1"]  # 读取名字为Sheet1的表

Total_time = 1920  # 总采样时间
time = 60  # 单次采样时间

for i in range(int(Total_time/time)):
    '''定义空数组'''
    concentration_1 = []

    '''分配数据'''
    row_num = 1
    while row_num < 61:
        concentration_1.append(sheet.cell(row=i * time + row_num, column=1).value)
        row_num = row_num + 1

    sensor_1 = numda * calculate_mean_value(concentration_1) + (1 - numda) * calculate_P(concentration_1) * calculate_Bout(
            concentration_1) * 6

    SC_1.append(sensor_1)

print(SC_1)

'''写出插值后的数据'''
np.savetxt('./data/移动机器人采样-第四组-10hz-7.15-2.05/MSC_data.txt', SC_1, fmt="%.2f", delimiter=' ')